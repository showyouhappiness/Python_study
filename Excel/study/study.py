# 导入文件
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import DataBarRule

# 创建空工作簿
wb = Workbook()
ws = wb.active

df = pd.read_excel(r'../sale_data.xlsx')


# 构造同时获取不同指标的函数
def get_data(date):
    create_cnt = df[df['创建日期'] == date]['order_id'].count()
    pay_cnt = df[df['付款日期'] == date]['order_id'].count()
    receive_cnt = df[df['收货日期'] == date]['order_id'].count()
    return_cnt = df[df['退款日期'] == date]['order_id'].count()
    return create_cnt, pay_cnt, receive_cnt, return_cnt


# 假设当日是2021-04-11
# 获取不同时间段的各指标值
df_view = pd.DataFrame([get_data('2021-04-11')
                           , get_data('2021-04-10')
                           , get_data('2021-04-04')]
                       , columns=['创建订单量', '付款订单量', '收货订单量', '退款订单量']
                       , index=['当日', '昨日', '上周同期']).T

df_view['环比'] = df_view['当日'] / df_view['昨日'] - 1
df_view['同比'] = df_view['当日'] / df_view['上周同期'] - 1
# 将DataFrame格式数据转化为openpyxl格式
for r in dataframe_to_rows(df_view, index=True, header=True):
    ws.append(r)

# 第二行是空的，删除第二行
ws.delete_rows(2)

# 给A1单元格进行赋值
ws['A1'] = '指标'

# 插入一行作为标题行
ws.insert_rows(1)
ws['A1'] = '电商业务方向 2021/4/11 日报'

# 将标题行的单元格进行合并
ws.merge_cells('A1:F1')  # 合并单元格

# 对第1行至第6行的单元格进行格式设置
for row in ws[1:6]:
    for c in row:
        # 字体设置
        c.font = Font(name='微软雅黑', size=12)
        # 对齐方式设置
        c.alignment = Alignment(horizontal="center")
        # 边框线设置
        c.border = Border(left=Side(border_style="thin", color="FF000000"),
                          right=Side(border_style="thin", color="FF000000"),
                          top=Side(border_style="thin", color="FF000000"),
                          bottom=Side(border_style="thin", color="FF000000"))

# 对标题行和表头行进行特殊设置
for row in ws[1:2]:
    for c in row:
        c.font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFFFF")
        c.fill = PatternFill(fill_type='solid', start_color='FFFF6100')

# 将环比和同比设置成百分比格式
for col in ws["E":"F"]:
    for r in col:
        r.number_format = '0.00%'

# 调整列宽
ws.column_dimensions['A'].width = 13
ws.column_dimensions['E'].width = 10

# 保存调整后的文件
wb.save(r'../核心指标.xlsx')

df_province = pd.DataFrame(df[df['创建日期'] == '2021-04-11'].groupby('省份')['order_id'].count())
df_province = df_province.reset_index()
df_province = df_province.sort_values(by='order_id', ascending=False)
df_province = df_province.rename(columns={'order_id': '创建订单量'})

# 设置进度条条件格式
rule = DataBarRule(start_type='min', end_type='max',
                   color="FF638EC6", showValue=True, minLength=None, maxLength=None)
ws.conditional_formatting.add('B1:B11', rule)

# 对第1行标题行进行设置
for c in ws[1]:
    c.font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFFFF")
    c.fill = PatternFill(fill_type='solid', start_color='FFFF6100')

# 调整列宽
ws.column_dimensions['A'].width = 17
ws.column_dimensions['B'].width = 13

# 保存调整后的文件
wb.save(r'../各省份销量情况.xlsx')
